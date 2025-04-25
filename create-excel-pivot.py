import pandas as pd
from openpyxl import load_workbook
from openpyxl.pivot.table import PivotTable, PivotTableStyleInfo
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.utils import get_column_letter

# 1. Create a Sample DataFrame
data = {
    'Region': ['North', 'South', 'East', 'West', 'North', 'South', 'East', 'West', 'North', 'South'],
    'Category': ['A', 'B', 'A', 'C', 'B', 'A', 'C', 'A', 'C', 'B'],
    'Product': ['P1', 'P2', 'P1', 'P3', 'P2', 'P1', 'P3', 'P1', 'P3', 'P2'],
    'Sales': [100, 150, 120, 200, 90, 110, 210, 130, 220, 160],
    'Quantity': [10, 15, 12, 20, 9, 11, 21, 13, 22, 16]
}
df = pd.DataFrame(data)

# 2. Define File and Sheet Names
excel_filename = 'data_with_pivot.xlsx'
data_sheet_name = 'DataSource'
pivot_sheet_name = 'SalesPivotTable'

print(f"Writing DataFrame to '{excel_filename}', sheet '{data_sheet_name}'...")

# 3. Use pd.ExcelWriter to write data and keep file open for openpyxl
# This automatically handles saving at the end of the 'with' block
try:
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Write the DataFrame to the 'DataSource' sheet
        df.to_excel(writer, sheet_name=data_sheet_name, index=False)

        # Get the openpyxl workbook and worksheet objects from the writer
        workbook = writer.book
        data_sheet = writer.sheets[data_sheet_name]

        print(f"Creating pivot table on sheet '{pivot_sheet_name}'...")

        # 4. Create the sheet for the pivot table
        pivot_sheet = workbook.create_sheet(title=pivot_sheet_name)

        # 5. Define the range of the source data (including headers)
        max_row = data_sheet.max_row
        max_col = data_sheet.max_column
        data_range_ref = CellRange(min_col=1, min_row=1, max_col=max_col, max_row=max_row)
        # Alternative using Reference (less direct for full range):
        # from openpyxl.worksheet.dimensions import DimensionHolder, Dimension
        # data_range_ref = Reference(data_sheet, min_col=1, min_row=1, max_col=max_col, max_row=max_row)

        # 6. Define Pivot Table Rows
        # References point to the *header* cells (Row 1) of the desired row fields
        pivot_rows = []
        # Add 'Region' (Column 1) as a row field
        pivot_rows.append(CellRange(min_col=1, min_row=1, max_col=1, max_row=1)) # Reference to A1
        # Add 'Category' (Column 2) as another row field
        pivot_rows.append(CellRange(min_col=2, min_row=1, max_col=2, max_row=1)) # Reference to B1

        # 7. Define Pivot Table Data/Values
        # References point to the *data* cells (Row 2 onwards) for aggregation
        pivot_data = []
        # Add 'Sales' (Column 4) to be summed
        sales_col_letter = get_column_letter(df.columns.get_loc('Sales') + 1) # +1 for 1-based index
        sales_range_ref = CellRange(f'{sales_col_letter}2:{sales_col_letter}{max_row}')
        pivot_data.append((sales_range_ref, 'Sum of Sales', 'sum'))

        # Add 'Quantity' (Column 5) to be summed
        qty_col_letter = get_column_letter(df.columns.get_loc('Quantity') + 1)
        qty_range_ref = CellRange(f'{qty_col_letter}2:{qty_col_letter}{max_row}')
        pivot_data.append((qty_range_ref, 'Total Quantity', 'sum'))

        # Add 'Sales' (Column 4) to be averaged
        # Using the same range ref as sum, but changing name and function
        pivot_data.append((sales_range_ref, 'Average Sales', 'average'))


        # 8. Create the PivotTable object
        # Note: openpyxl handles PivotCache creation automatically if location is None
        pt = PivotTable(name="SalesPivot", location=None) # location=None lets openpyxl manage cache

        # Assign the defined components
        pt.data_source = data_range_ref # Where the raw data lives
        pt.rows = pivot_rows             # Fields to use as rows
        # pt.cols = []                   # Define column fields here if needed
        pt.data = pivot_data             # Fields to aggregate in the values area

        # Optional: Apply a style
        pt.style = PivotTableStyleInfo(name="PivotStyleMedium9", showRowHeaders=True,
                                        showColHeaders=True, showRowStripes=False,
                                        showColStripes=False, firstHeaderRow=True,
                                        firstDataRow=False)

        # 9. Add the pivot table to the desired sheet ('PivotSheet') at cell 'A1'
        pivot_sheet.add_pivot_table(pt, "A1")

        print(f"Pivot table '{pt.name}' created successfully.")

    print(f"Workbook '{excel_filename}' saved successfully.")

except Exception as e:
    print(f"An error occurred: {e}")
